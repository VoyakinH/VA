# Наилучшее среднеквадратичное приближние.
# Воякин Алексей Янович ИУ7-44Б.

import openpyxl as xls
import numpy as np

# Задание функции.
def func(x, k):
    return x ** k

def solve_gauss(mat):
    n = len(mat)
    for k in range(n):
        for i in range(k + 1, n):
            coeff = -(mat[i][k] / mat[k][k])
            for j in range(k, n + 1):
                mat[i][j] += coeff * mat[k][j]
    a = [0 for i in range(n)]
    for i in range(n - 1, -1, -1):
        for j in range(n - 1, i, -1):
            mat[i][n] -= a[j] * mat[i][j]
        a[i] = mat[i][n] / mat[i][i]
    return a

def fill_slay(table, n):
    N = len(table)
    mat = [[0 for i in range(0, n + 1)] for j in range(0, n + 1)]
    col = [0 for i in range(0, n + 1)]
    for m in range(0, n + 1):
        for i in range(0, N):
            buff = table[i][2] * func(table[i][0], m)
            for k in range(0, n + 1):
                mat[m][k] += buff * func(table[i][0], k)
            col[m] += buff * table[i][1]
    for i in range(len(col)):
        mat[i].append(col[i])
    return mat

def count_coeffs(table, n):
    mat = fill_slay(table, n)
    a_arr = solve_gauss(mat)
    return a_arr

def main():
    i = 2
    nodes = xls.load_workbook('nodes.xlsx').active
    table = []
    while nodes.cell(row=i, column=1).value is not None:
        table.append([nodes.cell(row=i, column=1).value, nodes.cell(row=i, column=2).value, nodes.cell(row=i, column=3)])
        i += 1
    n = int(input('Введите степень полинома: '))
    a_arr = count_coeffs(table, n)

if __name__ == '__main__':
    main()