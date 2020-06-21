# Наилучшее среднеквадратичное приближние.
# Воякин Алексей Янович ИУ7-44Б.

import openpyxl as xls
import matplotlib.pyplot as plt

def fi(x, k):
    return x ** k

def solve_gauss(mat):
    n = len(mat)
    for k in range(n):
        for i in range(k + 1, n):
            coeff = -(mat[i][k] / mat[k][k])
            for j in range(k, n + 1):
                mat[i][j] += coeff * mat[k][j]
    a = [0 for i in range(n)]
    for i in range(n - 1, -1, -1):
        for j in range(n - 1, i, -1):
            mat[i][n] -= a[j] * mat[i][j]
        a[i] = mat[i][n] / mat[i][i]
    return a

def fill_slay(table, n):
    N = len(table)
    mat = [[0 for i in range(0, n + 1)] for j in range(0, n + 1)]
    col = [0 for i in range(0, n + 1)]
    for m in range(0, n + 1):
        for i in range(0, N):
            buff = table[i][2] * fi(table[i][0], m)
            for k in range(0, n + 1):
                mat[m][k] += buff * fi(table[i][0], k)
            col[m] += buff * table[i][1]
    for i in range(len(col)):
        mat[i].append(col[i])
    return mat
def count_coeffs(table, n):
    mat = fill_slay(table, n)
    a_arr = solve_gauss(mat)
    return a_arr

def draw_graph(table, n):
    dx = (table[-1][0] - table[0][0]) / 10
    fig, ax = plt.subplots()
    for i in range(1, n + 1):
        a_arr = count_coeffs(table, i)
        y = []
        x = []
        j = table[0][0] - dx
        while j <= table[-1][0] + dx:
            buff = 0
            for k in range(0, i + 1):
                buff += fi(j, k) * a_arr[k]
            y.append(buff)
            x.append(j)
            j += 0.01
        ax.plot(x, y, label='n = ' + str(i))
    ax.plot([a[0] for a in table], [a[1] for a in table], 'o', label='Date')
    plt.title("Наилучшее среднеквадратичное приближние.")
    plt.legend()
    plt.grid(True)
    plt.ylabel("Y")
    plt.xlabel("X")
    plt.show()

def main():
    i = 2
    nodes = xls.load_workbook('nodes.xlsx').active
    table = []
    while nodes.cell(row=i, column=1).value is not None:
        table.append([nodes.cell(row=i, column=1).value, nodes.cell(row=i, column=2).value,
                      nodes.cell(row=i, column=3).value])
        i += 1
    n = int(input('Введите максимальную степень полинома: '))
    draw_graph(table, n)

if __name__ == '__main__':
    main()
