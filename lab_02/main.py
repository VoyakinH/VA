# Интерполяция таблично заданных функций полиномом Ньютона.
# Воякин Алексей Янович ИУ7-44Б.

import openpyxl as xls
from math import sqrt
import numpy as np

# Задание функции.
def func(arg_x, arg_y):
    return pow(x, 2) + pow(y, 2)

# Нахождение необходимых начального и конечного индексов таблицы.
def find_indexes(data, power, arg):
    index = 0
    while arg > data[index]:
        index += 1
    ind_min = index - power // 2 - 1
    ind_max = index + (power // 2) + (power % 2) - 1
    if ind_max > len(data) - 1:
        ind_min -= ind_max - len(data) + 1
        ind_max = len(data) - 1
    elif ind_min < 0:
        ind_max += -ind_min
        ind_min = 0
    return ind_min+1, ind_max+1

# Вычисление разделенных разностей
def razd_razn(data, ind_start, ind_end):
    if ind_end - ind_start > 1:
        result = (razd_razn(data, ind_start, ind_end-1) - razd_razn(data, ind_start+1, ind_end))\
               / (data[ind_start][0] - data[ind_end][0])
    else:
        result = (data[ind_start][1] - data[ind_end][1]) / (data[ind_start][0] - data[ind_end][0])
    if ind_start == 0:
        razd_razn_founded.append(result)
    return result

# Нахождение значения функции
def find_root(data, arg):
    result = data[0][1]
    mul_of_diff = 1
    for i in range(0, len(data) - 1):
        mul_of_diff *= arg - data[i][0]
        result += mul_of_diff * razd_razn_founded[i]
    return result

# Функция интерполяции
def interpolation(data, arg_x, arg_y, pow_x, pow_y):
    data = (data[:,data[0,:].argsort()])[data[:, 0].argsort(), :]
    # Проверка на необходимое кол-во узлов.
    if pow_x + 1 > data.shape[0] - 1 or pow_y + 1 > data.shape[1] - 1:
        print('Недостаточное кол-во узлов для интерполяции.')
        exit(1)
    # Исключение экстраполяции
    if arg_x > data[len(data) - 1][0] or arg_x < data[1][0] or arg_y > data[0][len(data) - 1] or arg_y < data[0][1]:
        print('Экстраполяция недоступна.')
        exit(2)
    indexes_x = find_indexes(data[1:,0], pow_x, arg_x)
    indexes_y = find_indexes(data[0,1:], pow_y, arg_y)
    data = np.vstack([data[0], data[indexes_x[0]:indexes_x[1]+1]])
    data = np.column_stack((data[:, 0], data[:, indexes_y[0]:indexes_y[1] + 1]))
    buffer = []
    for i in range(0, len(data[:, 0]) - 1):
        if (pow_y > 0):
            razd_razn_founded.clear()
            razd_razn(np.transpose(np.vstack([data[0, 1:], data[i + 1, 1:]])), 0, len(data[0]) - 2)
        buffer.append(find_root(np.transpose(np.vstack([data[0, 1:], data[i + 1, 1:]])), arg_y))
    if (pow_x > 0):
        razd_razn_founded.clear()
        razd_razn(np.column_stack((data[1:, 0], np.transpose(buffer))), 0, len(buffer) - 1)
    return find_root(np.column_stack((data[1:, 0], np.transpose(buffer))), arg_x)

# Чтение таблицы точек из файла.
i = 1; j = 1
nodes = xls.load_workbook('nodes.xlsx', data_only=True).active
row = []
razd_razn_founded = []
while nodes.cell(row=i, column=1).value is not None:
    while nodes.cell(row=1, column=j).value is not None:
        try:
            row.append(float(nodes.cell(row=i, column=j).value))
        except:
            row.append(-pow(10, 6))
        j += 1
    i += 1; j = 1
    try:
        table = np.vstack([table, row])
    except:
        table = np.vstack([row])
    row.clear()

# Чтение X и Y с клавиатуры.
x, y = map(float, input('Введите X, Y через пробел: ').split())
# Чтение степеней полиномов с клавиатуры.
nx, ny = map(int, input('Введите степени полиномов Nx, Ny через пробел: ').split())

# Выполнение алгоритмов.
is_known = False
i, = np.where(table[:, 0] == x); j, = np.where(table[0] == y)
if len(i) != 0 and len(j) != 0:
    print('\nЗначение функции есть в таблице: F(', x, ', ', y, ') = ', table[i[0], j[0]], sep='')
    is_known = True
if not is_known:
    founded_root = float("%.7f" % interpolation(table, x, y, nx, ny))
    exact_root = float("%.7f" % func(x, y))
    print("\nНайденное значение: ", founded_root)
    print("Точное значение:    ", exact_root)
    print("Относительная ошибка: %.2f" % abs(abs(exact_root - founded_root) / exact_root * 100), '%')